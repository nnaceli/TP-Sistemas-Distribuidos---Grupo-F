from flask import Blueprint, request, jsonify, Response
from openpyxl import Workbook
import io
import requests
from grpc import RpcError 
# from GrpcService.GrpcDonacionService import listar_las_donaciones # Eliminamos esta importación

# URL base del servidor REST de Java (informe-api-server)
# Asumo que el puerto por defecto de Spring Boot es 8080. AJUSTA SI ES NECESARIO.
JAVA_REST_BASE_URL = "http://localhost:8080" 

class Donacion:
    """Clase simple para simular el objeto de donación completo que devuelve Java."""
    def __init__(self, data):
        self.fechaAlta = data.get('fechaAlta', '')
        self.descripcion = data.get('descripcion', '')
        self.cantidad = data.get('cantidad', 0)
        self.eliminado = data.get('eliminado', False)
        # Campos de usuario y modificación completos
        self.usuarioAlta = data.get('usuarioAlta', '')
        self.usuarioModificacion = data.get('usuarioModificacion', '')
        self.categoria = data.get('categoria', 'OTRAS')

def listar_las_donaciones_completas(token):
    """
    Llama al servicio REST completo en el servidor Java para obtener todas las donaciones.
    El Token se pasa para la autenticación en el servidor Java.
    """
    endpoint = f"{JAVA_REST_BASE_URL}/api/donaciones/listar-todas"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    
    # Intenta hacer la petición
    response = requests.get(endpoint, headers=headers)
    response.raise_for_status() # Lanza un error HTTP si la respuesta no es 2xx

    data = response.json()
    
    # Convertimos los diccionarios JSON en objetos Donacion para mantener la lógica original del código
    donaciones_completas = [Donacion(item) for item in data]
    return donaciones_completas


def extraer_token():
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return None
    return auth_header.split(" ")[1]

informe_excel_donaciones_bp = Blueprint('informes_donaciones', __name__)

@informe_excel_donaciones_bp.route('/excel', methods=['GET'])
def generar_informe_excel():
    try:
        token = extraer_token()
        if not token:
            return jsonify({"error": "Token no proporcionado"}), 401

        donaciones = listar_las_donaciones_completas(token)

        # agrupacion de donaciones por categorias
        donaciones_agrupadas = {
            'ROPA': [],
            'ALIMENTOS': [],
            'JUGUETES': [],
            'UTILES ESCOLARES': []
        }
        
        # 🚨 NOTA: Asegúrate de que el servidor Java devuelva la categoría en mayúsculas 
        # o ajusta esta lógica de conversión.
        for donacion in donaciones:
            # Aseguramos que la categoría sea un string antes de intentar .upper()
            categoria = str(donacion.categoria).upper()
            if categoria in donaciones_agrupadas:
                donaciones_agrupadas[categoria].append(donacion)
            else:
                # Opcional: Manejar categorías no mapeadas
                pass
        
        #creacion de archivo excel y escritura de datos
        wb = Workbook()
        
        # se elimina hoja creada por defecto
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        headers = ['Fecha de Alta', 'Descripción', 'Cantidad', 'Eliminado', 
                   'Usuario Alta', 'Usuario Modificación'] # Corregidos los headers para incluir los nuevos campos

        for categoria, lista_donaciones in donaciones_agrupadas.items():
            
            # Crea la hoja con el nombre de la Categoría
            ws = wb.create_sheet(title=categoria)
            ws.append(headers) 
            
            # Vuelca los datos de las donaciones en esa hoja
            for donacion in lista_donaciones:
                ws.append([
                    donacion.fechaAlta, 
                    donacion.descripcion, 
                    donacion.cantidad, 
                    donacion.eliminado, 
                    # 🚨 CAMBIO: Incluimos los campos de usuario y modificación que faltaban antes
                    donacion.usuarioAlta, 
                    donacion.usuarioModificacion,
                ])
                
        # se guarda el archivo y se envía para descargar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = "Informe_Donaciones.xlsx"
        
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )

    except requests.exceptions.RequestException as e:
        # Captura errores de conexión o HTTP (4xx, 5xx) del servidor Java
        error_msg = f"Error de comunicación con el servidor Java ({JAVA_REST_BASE_URL}): {str(e)}"
        return jsonify({"error": error_msg}), 503 # Servicio no disponible
    
    except Exception as e:
        # Otros errores (por ejemplo, procesamiento del Excel)
        return jsonify({"error": f"Error interno al generar el informe: {str(e)}"}), 500

# Esta parte solo se usa si ejecutas este archivo directamente como servidor de pruebas.
if __name__ == '__main__':
    from flask import Flask
    app = Flask(__name__)
    # Se utiliza un prefijo de URL para la ruta
    app.register_blueprint(informe_excel_donaciones_bp, url_prefix='/api/informes/donaciones') 
    app.run(port=5000, debug=True)
